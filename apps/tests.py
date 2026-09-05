import pytest
import json
from django.test import TestCase, Client, override_settings
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from unittest.mock import patch, MagicMock

User = get_user_model()

# ========================
# AUTHENTICATION & ACCOUNTS
# ========================

class TestRegistration(TestCase):
    def setUp(self):
        self.client = Client()
        self.register_url = reverse('accounts:register')

    def test_registration_page_loads(self):
        response = self.client.get(self.register_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'accounts/register.html')

    def test_user_can_register(self):
        data = {
            'username': 'testuser',
            'email': 'test@example.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'student_id': '2021-001-001',
            'department': 'cse',
            'phone': '01700000000',
        }
        response = self.client.post(self.register_url, data)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(User.objects.filter(username='testuser').exists())
        user = User.objects.get(username='testuser')
        self.assertFalse(user.is_membership_paid)

    def test_registration_requires_unique_username(self):
        User.objects.create_user(username='testuser', password='test123', is_membership_paid=True, email_verified=True)
        data = {
            'username': 'testuser',
            'email': 'test2@example.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
        }
        response = self.client.post(self.register_url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'already exists')

    def test_registration_requires_matching_passwords(self):
        data = {
            'username': 'testuser2',
            'email': 'test2@example.com',
            'password1': 'ComplexPass123!',
            'password2': 'DifferentPass456!',
        }
        response = self.client.post(self.register_url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'match')

    def test_blank_student_id_allowed(self):
        data = {
            'username': 'testuser3',
            'email': 'test3@example.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'student_id': '',
            'department': 'cse',
        }
        response = self.client.post(self.register_url, data)
        self.assertEqual(response.status_code, 302)


class TestLogin(TestCase):
    def setUp(self):
        self.client = Client()
        self.login_url = reverse('accounts:login')
        self.user = User.objects.create_user(username='testuser', password='testpass123', email='test@example.com', is_membership_paid=True, email_verified=True)

    def test_login_page_loads(self):
        response = self.client.get(self.login_url)
        self.assertEqual(response.status_code, 200)

    def test_user_can_login(self):
        response = self.client.post(self.login_url, {'username': 'testuser', 'password': 'testpass123'})
        self.assertEqual(response.status_code, 302)

    def test_login_with_invalid_credentials_fails(self):
        response = self.client.post(self.login_url, {'username': 'testuser', 'password': 'wrongpass'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Please enter')

    def test_suspended_user_cannot_login(self):
        self.user.is_suspended = True
        self.user.save()
        response = self.client.post(self.login_url, {'username': 'testuser', 'password': 'testpass123'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'suspended')


class TestPermissions(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='regular', password='testpass', is_membership_paid=True, email_verified=True)
        self.admin = User.objects.create_superuser(username='admin', email='admin@test.com', password='adminpass')
        self.admin.role = 'admin'
        self.admin.save()

    def test_unauthenticated_redirected_to_login(self):
        protected_urls = ['dashboard:home', 'accounts:profile', 'accounts:edit_profile']
        for url_name in protected_urls:
            response = self.client.get(reverse(url_name))
            self.assertEqual(response.status_code, 302)

    def test_admin_dashboard_requires_admin(self):
        self.client.login(username='regular', password='testpass')
        response = self.client.get(reverse('dashboard:admin_home'))
        self.assertEqual(response.status_code, 302)

    def test_admin_dashboard_allows_admin(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_home'))
        self.assertEqual(response.status_code, 200)


# ========================
# POST LIFECYCLE
# ========================

class TestPostLifecycle(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='poster', password='testpass', is_membership_paid=True, email_verified=True)
        from apps.membership.models import Membership, MembershipPlan
        plan = MembershipPlan.objects.create(name='Annual Membership', price=100, duration_days=365)
        Membership.objects.create(user=self.user, plan=plan, is_active=True,
                                  started_at=timezone.now(),
                                  expires_at=timezone.now() + timedelta(days=365))
        self.client.login(username='poster', password='testpass')
        from apps.posts.models import Category, CampusLocation
        self.category = Category.objects.create(name='Electronics', slug='electronics')
        self.location = CampusLocation.objects.create(name='Building A', slug='building-a')
        self.create_url = reverse('posts:create')

    def test_create_page_loads(self):
        response = self.client.get(self.create_url)
        self.assertEqual(response.status_code, 200)

    def test_user_can_create_lost_post(self):
        data = {
            'title': 'Lost iPhone',
            'description': 'Black iPhone 14 in Building A',
            'post_type': 'lost',
            'category': self.category.pk,
            'location': self.location.pk,
            'date_lost_found': timezone.now().date(),
        }
        response = self.client.post(self.create_url, data)
        self.assertEqual(response.status_code, 302)
        from apps.posts.models import Post
        self.assertTrue(Post.objects.filter(title='Lost iPhone').exists())

    def test_user_can_create_found_post(self):
        data = {
            'title': 'Found Keys',
            'description': 'Found a set of keys near the library',
            'post_type': 'found',
            'category': self.category.pk,
            'location': self.location.pk,
            'date_lost_found': timezone.now().date(),
        }
        response = self.client.post(self.create_url, data)
        self.assertEqual(response.status_code, 302)
        from apps.posts.models import Post
        self.assertTrue(Post.objects.filter(title='Found Keys').exists())

    def test_user_cannot_create_post_with_blank_title(self):
        data = {
            'title': '',
            'description': 'Some description',
            'post_type': 'lost',
        }
        response = self.client.post(self.create_url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'required')

    def test_user_can_edit_own_post(self):
        from apps.posts.models import Post
        post = Post.objects.create(
            user=self.user, title='Original Title', description='Test',
            post_type='lost', date_lost_found=timezone.now().date()
        )
        edit_url = reverse('posts:edit', kwargs={'pk': post.pk})
        response = self.client.post(edit_url, {
            'title': 'Updated Title',
            'description': 'Updated description',
            'post_type': 'lost',
            'category': self.category.pk,
            'location': self.location.pk,
            'date_lost_found': timezone.now().date(),
        })
        self.assertEqual(response.status_code, 302)
        post.refresh_from_db()
        self.assertEqual(post.title, 'Updated Title')

    def test_user_can_delete_own_post(self):
        from apps.posts.models import Post
        post = Post.objects.create(
            user=self.user, title='To Delete', description='Test',
            post_type='lost', date_lost_found=timezone.now().date()
        )
        delete_url = reverse('posts:delete', kwargs={'pk': post.pk})
        response = self.client.post(delete_url)
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Post.objects.filter(pk=post.pk).exists())

    def test_user_can_mark_resolved(self):
        from apps.posts.models import Post
        post = Post.objects.create(
            user=self.user, title='Resolve Me', description='Test',
            post_type='lost', date_lost_found=timezone.now().date()
        )
        resolve_url = reverse('posts:resolve', kwargs={'pk': post.pk})
        response = self.client.post(resolve_url)
        self.assertEqual(response.status_code, 302)
        post.refresh_from_db()
        self.assertEqual(post.status, 'resolved')
        self.assertTrue(post.is_resolved)

    def test_user_cannot_edit_anothers_post(self):
        other_user = User.objects.create_user(username='other', password='testpass', is_membership_paid=True, email_verified=True)
        from apps.posts.models import Post
        post = Post.objects.create(
            user=other_user, title='Others Post', description='Test',
            post_type='lost', date_lost_found=timezone.now().date()
        )
        edit_url = reverse('posts:edit', kwargs={'pk': post.pk})
        response = self.client.get(edit_url)
        self.assertEqual(response.status_code, 404)


# ========================
# BROWSE & SEARCH
# ========================

class TestBrowseAndSearch(TestCase):
    def setUp(self):
        self.client = Client()
        from apps.posts.models import Post, Category, CampusLocation
        self.cat = Category.objects.create(name='Books', slug='books')
        self.loc = CampusLocation.objects.create(name='Library', slug='library')
        self.user = User.objects.create_user(username='browser', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='browser', password='testpass')
        for i in range(15):
            Post.objects.create(
                user=self.user, title=f'Test Post {i}', description=f'Description {i}',
                post_type='lost' if i % 2 == 0 else 'found',
                category=self.cat if i < 10 else None,
                location=self.loc if i < 10 else None,
                date_lost_found=timezone.now().date(),
            )

    def test_browse_page_loads(self):
        response = self.client.get(reverse('posts:browse'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Test Post')

    def test_browse_filters_by_type(self):
        response = self.client.get(reverse('posts:browse') + '?type=lost')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '>Test Post 0</h3>')
        self.assertNotContains(response, '>Test Post 1</h3>')

    def test_browse_type_filter_options_for_users(self):
        response = self.client.get(reverse('posts:browse'))
        self.assertContains(response, '<option value="">All Types</option>')
        self.assertContains(response, 'value="lost"')
        self.assertContains(response, 'value="found"')

    def test_user_sidebar_hides_lost_found_items_links(self):
        response = self.client.get(reverse('posts:browse'))
        self.assertNotContains(response, '/browse/?type=lost')
        self.assertNotContains(response, '/browse/?type=found')

    def test_browse_has_pagination(self):
        response = self.client.get(reverse('posts:browse'))
        self.assertContains(response, 'page')

    def test_post_detail_shows(self):
        from apps.posts.models import Post
        post = Post.objects.first()
        response = self.client.get(reverse('posts:detail', kwargs={'pk': post.pk}))
        self.assertEqual(response.status_code, 200)


# ========================
# MATCHING ENGINE
# ========================

class TestMatchingEngine(TestCase):
    def setUp(self):
        from apps.posts.models import Post, Category, CampusLocation, PostTag
        self.user = User.objects.create_user(username='matcher', password='testpass', is_membership_paid=True, email_verified=True)
        self.other_user = User.objects.create_user(username='other_matcher', password='testpass', is_membership_paid=True, email_verified=True)
        self.cat = Category.objects.create(name='Electronics', slug='electronics')
        self.loc = CampusLocation.objects.create(name='Campus A', slug='campus-a')
        self.lost_post = Post.objects.create(
            user=self.user, title='Lost Phone', description='Black Samsung phone',
            post_type='lost', category=self.cat, location=self.loc,
            date_lost_found=timezone.now().date(),
        )
        self.found_post = Post.objects.create(
            user=self.other_user, title='Found Phone', description='Found a black Samsung',
            post_type='found', category=self.cat, location=self.loc,
            date_lost_found=timezone.now().date(),
        )
        PostTag.objects.create(post=self.lost_post, name='phone')
        PostTag.objects.create(post=self.found_post, name='phone')

    @patch('apps.ai_engine.utils.generate_embedding')
    @patch('apps.ai_engine.utils._ranked_candidates')
    def test_match_created_on_post_creation(self, mock_candidates, mock_embedding):
        mock_embedding.return_value = [0.1] * 256
        mock_candidates.return_value = [(self.found_post, 0.85)]
        from apps.ai_engine.utils import find_matches_for_post
        matches = find_matches_for_post(self.lost_post)
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0].matched_post, self.found_post)
        # 0.60*0.85 (semantic) + 0.15 (category) + 0.10 (location) + 0.10 (date) + 0.05 (tags) = 0.91
        self.assertAlmostEqual(matches[0].similarity_score, 0.91, places=2)
        self.assertEqual(matches[0].match_strength, 'strong')

    def test_hybrid_scoring_prefers_structured_matches(self):
        from apps.ai_engine.utils import hybrid_match_score
        from apps.posts.models import Post
        other_cat_post = Post.objects.create(
            user=self.user, title='Found Wallet', description='Leather wallet',
            post_type='found', date_lost_found=timezone.now().date() - timedelta(days=30),
        )
        same_score, same_meta = hybrid_match_score(self.lost_post, self.found_post, 0.5)
        diff_score, diff_meta = hybrid_match_score(self.lost_post, other_cat_post, 0.5)
        self.assertGreater(same_score, diff_score)
        raw_score, raw_meta = hybrid_match_score(self.lost_post, other_cat_post, 1.0)
        self.assertAlmostEqual(raw_score, 0.60, places=2)

    def test_match_suggestion_page_requires_login(self):
        response = self.client.get(reverse('ai:matches'))
        self.assertEqual(response.status_code, 302)

    def test_match_suggestion_page_works(self):
        from apps.ai_engine.models import MatchSuggestion
        MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='matcher', password='testpass')
        response = self.client.get(reverse('ai:matches'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Lost Phone')

    def test_dismiss_match_post(self):
        from apps.ai_engine.models import MatchSuggestion
        match = MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='matcher', password='testpass')
        response = self.client.post(reverse('ai:dismiss_match', kwargs={'match_id': match.pk}))
        self.assertEqual(response.status_code, 302)
        match.refresh_from_db()
        self.assertEqual(match.status, 'dismissed')

    def test_accept_match_post(self):
        from apps.ai_engine.models import MatchSuggestion
        match = MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='matcher', password='testpass')
        response = self.client.post(reverse('ai:accept_match', kwargs={'match_id': match.pk}))
        self.assertEqual(response.status_code, 302)
        match.refresh_from_db()
        self.assertEqual(match.status, 'accepted')
        self.assertTrue(match.is_accepted)

    def test_dismiss_match_get_rejected(self):
        from apps.ai_engine.models import MatchSuggestion
        match = MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='matcher', password='testpass')
        response = self.client.get(reverse('ai:dismiss_match', kwargs={'match_id': match.pk}))
        self.assertEqual(response.status_code, 302)
        match.refresh_from_db()
        self.assertEqual(match.status, 'pending')

    def test_accept_match_get_rejected(self):
        from apps.ai_engine.models import MatchSuggestion
        match = MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='matcher', password='testpass')
        response = self.client.get(reverse('ai:accept_match', kwargs={'match_id': match.pk}))
        self.assertEqual(response.status_code, 302)
        match.refresh_from_db()
        self.assertEqual(match.status, 'pending')

    def test_unauthorized_user_cannot_dismiss_match(self):
        from apps.ai_engine.models import MatchSuggestion
        stranger = User.objects.create_user(username='stranger', password='testpass', is_membership_paid=True, email_verified=True)
        match = MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='stranger', password='testpass')
        response = self.client.post(reverse('ai:dismiss_match', kwargs={'match_id': match.pk}))
        self.assertEqual(response.status_code, 302)
        match.refresh_from_db()
        self.assertEqual(match.status, 'pending')

    def test_unauthorized_user_cannot_accept_match(self):
        from apps.ai_engine.models import MatchSuggestion
        stranger = User.objects.create_user(username='stranger2', password='testpass', is_membership_paid=True, email_verified=True)
        match = MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        self.client.login(username='stranger2', password='testpass')
        response = self.client.post(reverse('ai:accept_match', kwargs={'match_id': match.pk}))
        self.assertEqual(response.status_code, 302)
        match.refresh_from_db()
        self.assertEqual(match.status, 'pending')

    def test_ranked_candidates_filters_by_opposite_type(self):
        from apps.ai_engine.utils import _ranked_candidates, generate_embedding, build_text_for_post
        from apps.posts.models import Post
        lost2 = Post.objects.create(
            user=self.other_user, title='Lost Wallet', description='Brown leather wallet',
            post_type='lost', date_lost_found=timezone.now().date(),
        )
        with patch('apps.ai_engine.utils.generate_embedding', return_value=[0.1] * 256):
            with patch('apps.ai_engine.utils.vector_backend_available', return_value=False):
                candidates = _ranked_candidates(self.lost_post, [0.1] * 256)
        candidate_pks = [p.pk for p, _ in candidates]
        self.assertIn(self.found_post.pk, candidate_pks)
        self.assertNotIn(lost2.pk, candidate_pks)

    def test_ranked_candidates_filters_found_excludes_found(self):
        from apps.ai_engine.utils import _ranked_candidates
        from apps.posts.models import Post
        found2 = Post.objects.create(
            user=self.other_user, title='Found Wallet', description='Brown leather wallet',
            post_type='found', date_lost_found=timezone.now().date(),
        )
        with patch('apps.ai_engine.utils.generate_embedding', return_value=[0.1] * 256):
            with patch('apps.ai_engine.utils.vector_backend_available', return_value=False):
                candidates = _ranked_candidates(self.found_post, [0.1] * 256)
        candidate_pks = [p.pk for p, _ in candidates]
        self.assertIn(self.lost_post.pk, candidate_pks)
        self.assertNotIn(found2.pk, candidate_pks)

    def test_match_below_threshold_not_created(self):
        from apps.ai_engine.models import MatchSuggestion
        from apps.posts.models import Post, Category
        diff_cat = Category.objects.create(name='Clothing', slug='clothing')
        diff_cat_post = Post.objects.create(
            user=self.other_user, title='Found Scarf', description='Red wool scarf',
            post_type='found', category=diff_cat,
            date_lost_found=timezone.now().date() - timedelta(days=30),
        )
        with patch('apps.ai_engine.utils.generate_embedding', return_value=[0.1] * 256):
            with patch('apps.ai_engine.utils._ranked_candidates', return_value=[(diff_cat_post, 0.1)]):
                from apps.ai_engine.utils import find_matches_for_post
                matches = find_matches_for_post(self.lost_post)
        self.assertEqual(len(matches), 0)
        self.assertEqual(MatchSuggestion.objects.count(), 0)

    def test_duplicate_suggestion_not_created(self):
        from apps.ai_engine.models import MatchSuggestion
        MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.9)
        with patch('apps.ai_engine.utils.generate_embedding', return_value=[0.1] * 256):
            with patch('apps.ai_engine.utils._ranked_candidates', return_value=[(self.found_post, 0.9)]):
                from apps.ai_engine.utils import find_matches_for_post
                find_matches_for_post(self.lost_post)
        self.assertEqual(MatchSuggestion.objects.count(), 1)

    def test_existing_suggestion_updated_not_duplicated(self):
        from apps.ai_engine.models import MatchSuggestion
        MatchSuggestion.objects.create(post=self.lost_post, matched_post=self.found_post, similarity_score=0.6)
        with patch('apps.ai_engine.utils.generate_embedding', return_value=[0.1] * 256):
            with patch('apps.ai_engine.utils._ranked_candidates', return_value=[(self.found_post, 0.95)]):
                from apps.ai_engine.utils import find_matches_for_post
                find_matches_for_post(self.lost_post)
        match = MatchSuggestion.objects.get(post=self.lost_post, matched_post=self.found_post)
        # 0.60*0.95 + 0.15 + 0.10 + 0.10 + 0.05 = 0.97
        self.assertAlmostEqual(match.similarity_score, 0.97, places=2)
        self.assertEqual(MatchSuggestion.objects.count(), 1)

    @patch('apps.ai_engine.utils.generate_embedding')
    def test_jina_failure_does_not_prevent_post_creation(self, mock_embedding):
        mock_embedding.return_value = None
        from apps.ai_engine.utils import find_matches_for_post
        matches = find_matches_for_post(self.lost_post)
        self.assertEqual(matches, [])

    def test_match_notification_created_for_both_users(self):
        from apps.ai_engine.models import MatchSuggestion
        from apps.notifications.models import Notification
        match = MatchSuggestion.objects.create(
            post=self.lost_post, matched_post=self.found_post, similarity_score=0.9
        )
        self.assertTrue(Notification.objects.filter(user=self.user, notification_type='match_found').exists())
        self.assertTrue(Notification.objects.filter(user=self.other_user, notification_type='match_found').exists())

    def test_match_strength_classifies_correctly(self):
        from apps.ai_engine.utils import _get_match_strength
        self.assertEqual(_get_match_strength(0.80), 'strong')
        self.assertEqual(_get_match_strength(0.60), 'possible')
        self.assertEqual(_get_match_strength(0.45), 'possible')

    def test_hybrid_match_score_returns_tuple(self):
        from apps.ai_engine.utils import hybrid_match_score
        result = hybrid_match_score(self.lost_post, self.found_post, 0.8)
        self.assertIsInstance(result, tuple)
        self.assertEqual(len(result), 2)
        score, meta = result
        self.assertGreaterEqual(score, 0.0)
        self.assertLessEqual(score, 1.0)
        self.assertGreaterEqual(meta, 0.0)
        self.assertLessEqual(meta, 1.0)


# ========================
# AI SEARCH
# ========================

class TestAISearch(TestCase):
    def setUp(self):
        from apps.posts.models import Post, Category, CampusLocation
        self.user = User.objects.create_user(username='searcher', password='testpass', is_membership_paid=True, email_verified=True)
        self.cat = Category.objects.create(name='Electronics', slug='electronics')
        self.loc = CampusLocation.objects.create(name='Library', slug='library')
        self.post = Post.objects.create(
            user=self.user, title='Lost black Samsung phone',
            description='Lost near the library entrance',
            post_type='lost', category=self.cat, location=self.loc,
            date_lost_found=timezone.now().date(),
        )

    def test_search_page_requires_login(self):
        response = self.client.get(reverse('ai:search'))
        self.assertEqual(response.status_code, 302)

    def test_search_page_loads(self):
        self.client.login(username='searcher', password='testpass')
        response = self.client.get(reverse('ai:search'))
        self.assertEqual(response.status_code, 200)

    @patch('apps.ai_engine.views.semantic_search_posts')
    @patch('apps.ai_engine.views.generate_embedding')
    @patch('apps.ai_engine.views.vector_backend_available', return_value=True)
    def test_semantic_search_used_when_available(self, mock_avail, mock_emb, mock_semantic):
        mock_emb.return_value = [0.1] * 256
        mock_semantic.return_value = [(self.post, 0.85)]
        self.client.login(username='searcher', password='testpass')
        response = self.client.get(reverse('ai:search') + '?q=black+samsung+phone')
        self.assertEqual(response.status_code, 200)
        mock_semantic.assert_called_once()

    @patch('apps.ai_engine.views.generate_embedding', return_value=None)
    @patch('apps.ai_engine.views.vector_backend_available', return_value=True)
    def test_keyword_fallback_when_embedding_fails(self, mock_avail, mock_emb):
        self.client.login(username='searcher', password='testpass')
        response = self.client.get(reverse('ai:search') + '?q=samsung')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Samsung')

    @patch('apps.ai_engine.views.vector_backend_available', return_value=False)
    def test_keyword_fallback_when_vector_backend_unavailable(self, mock_avail):
        self.client.login(username='searcher', password='testpass')
        response = self.client.get(reverse('ai:search') + '?q=samsung')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Samsung')

    def test_filters_still_work(self):
        self.client.login(username='searcher', password='testpass')
        response = self.client.get(reverse('ai:search') + '?type=lost&category=electronics')
        self.assertEqual(response.status_code, 200)

    def test_empty_search_returns_nothing(self):
        self.client.login(username='searcher', password='testpass')
        response = self.client.get(reverse('ai:search') + '?q=')
        self.assertEqual(response.status_code, 200)


# ========================
# EMBEDDING GENERATION
# ========================

class TestEmbeddingGeneration(TestCase):
    def setUp(self):
        from apps.posts.models import Post, Category
        self.user = User.objects.create_user(username='embedder', password='testpass', is_membership_paid=True, email_verified=True)
        self.cat = Category.objects.create(name='Documents', slug='documents')
        self.post = Post.objects.create(
            user=self.user, title='Lost student ID card',
            description='Blue student ID card with photo',
            post_type='lost', category=self.cat,
            date_lost_found=timezone.now().date(),
        )

    def test_build_text_includes_key_fields(self):
        from apps.ai_engine.utils import build_text_for_post
        text = build_text_for_post(self.post)
        self.assertIn('Lost', text)
        self.assertIn('student ID card', text)
        self.assertIn('Documents', text)

    def test_build_text_excludes_contact_info(self):
        self.post.contact_info = 'Phone: 01712345678'
        self.post.save()
        from apps.ai_engine.utils import build_text_for_post
        text = build_text_for_post(self.post)
        self.assertNotIn('01712345678', text)

    @patch('apps.ai_engine.utils.store_post_embedding')
    @patch('apps.ai_engine.utils.generate_embedding')
    def test_refresh_post_embedding(self, mock_embedding, mock_store):
        mock_embedding.return_value = [0.5] * 256
        mock_store.return_value = (True,)
        from apps.ai_engine.utils import refresh_post_embedding
        result = refresh_post_embedding(self.post)
        self.assertIsNotNone(result)
        mock_embedding.assert_called_once()

    @patch('apps.ai_engine.utils.generate_embedding', return_value=None)
    def test_refresh_returns_none_on_failure(self, mock_embedding):
        from apps.ai_engine.utils import refresh_post_embedding
        result = refresh_post_embedding(self.post)
        self.assertIsNone(result)

    def test_post_without_pk_skips_matching(self):
        from apps.ai_engine.utils import find_matches_for_post
        from apps.posts.models import Post
        unsaved = Post(user=self.user, title='test', description='test', post_type='lost', date_lost_found=timezone.now().date())
        matches = find_matches_for_post(unsaved)
        self.assertEqual(matches, [])


# ========================
# MEMBERSHIP SYSTEM
# ========================

class TestMembership(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='member', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='member', password='testpass')
        from apps.membership.models import MembershipPlan
        self.plan = MembershipPlan.objects.create(
            name='Premium Monthly', price=Decimal('50.00'), duration_days=30, is_active=True
        )

    def test_membership_page_loads(self):
        response = self.client.get(reverse('membership:index'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Annual Membership')

    def test_purchase_initiation_creates_membership(self):
        response = self.client.get(reverse('membership:purchase', kwargs={'plan_id': self.plan.pk}))
        self.assertEqual(response.status_code, 302)
        from apps.membership.models import Membership
        self.assertTrue(Membership.objects.filter(user=self.user).exists())

    def test_membership_success_page(self):
        from apps.membership.models import Membership
        Membership.objects.create(
            user=self.user, plan=self.plan, is_active=True,
            started_at=timezone.now(), expires_at=timezone.now() + timedelta(days=30)
        )
        response = self.client.get(reverse('membership:success'))
        self.assertEqual(response.status_code, 200)


# ========================
# PAYMENTS / SSLCOMMERZ
# ========================

class TestPayments(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='payer', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='payer', password='testpass')

    @patch('apps.payments.views.initiate_payment')
    def test_payment_initiation(self, mock_initiate):
        mock_initiate.return_value = 'http://mock-gateway-url.com'
        response = self.client.get(reverse('membership:index'))
        self.assertEqual(response.status_code, 200)

    @patch('apps.payments.views.verify_sslcommerz_payment')
    def test_payment_success_callback(self, mock_verify):
        from apps.payments.models import Payment
        from apps.membership.models import MembershipPlan, Membership
        plan = MembershipPlan.objects.create(name='Test', price=Decimal('100'), duration_days=30)
        payment = Payment.objects.create(
            user=self.user, transaction_id='TXN123', amount=Decimal('100.00'),
            payment_type='membership', status='pending', sslcommerz_tran_id='SSL-TXN-001',
        )
        mock_verify.return_value = {'status': 'VALID', 'bank_tran_id': 'BANK001'}
        response = self.client.post(reverse('payments:success'), {
            'tran_id': 'SSL-TXN-001',
            'status': 'VALID',
            'val_id': 'VAL001',
        })
        self.assertEqual(response.status_code, 302)
        payment.refresh_from_db()
        self.assertEqual(payment.status, 'completed')
        self.user.refresh_from_db()
        self.assertTrue(self.user.is_membership_paid)


# ========================
# ADMIN MODERATION
# ========================

class TestAdminModeration(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin = User.objects.create_superuser(username='admin', email='admin@test.com', password='adminpass')
        self.admin.role = 'admin'
        self.admin.save()
        self.client.login(username='admin', password='adminpass')

    def test_admin_dashboard_loads(self):
        response = self.client.get(reverse('dashboard:admin_home'))
        self.assertEqual(response.status_code, 200)

    def test_page_titles_match_sidebar(self):
        pages = [
            ('dashboard:admin_home', 'Dashboard'),
            ('dashboard:admin_users', 'Users'),
            ('posts:browse', 'Browse Posts'),
            ('notifications:list', 'Notifications'),
            ('dashboard:admin_memberships', 'Memberships'),
            ('dashboard:admin_revenue', 'Revenue'),
            ('accounts:settings', 'Settings'),
        ]
        for url_name, title in pages:
            with self.subTest(page=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, f'<title>{title} | IUBAT SmartFind</title>')
                self.assertContains(response, f'<h2 class="text-lg font-semibold text-gray-800 pl-1">{title}</h2>')

    def test_admin_users_page(self):
        response = self.client.get(reverse('dashboard:admin_users'))
        self.assertEqual(response.status_code, 200)

    def test_admin_posts_page(self):
        response = self.client.get(reverse('dashboard:admin_posts'))
        self.assertEqual(response.status_code, 200)

    def test_admin_categories_page(self):
        response = self.client.get(reverse('dashboard:admin_categories'))
        self.assertEqual(response.status_code, 200)

    def test_admin_locations_page(self):
        response = self.client.get(reverse('dashboard:admin_locations'))
        self.assertEqual(response.status_code, 200)

    def test_admin_revenue_page(self):
        response = self.client.get(reverse('dashboard:admin_revenue'))
        self.assertEqual(response.status_code, 200)

    def test_admin_reports_page(self):
        response = self.client.get(reverse('dashboard:admin_reports'))
        self.assertEqual(response.status_code, 200)

    def test_admin_analytics_page(self):
        response = self.client.get(reverse('dashboard:admin_analytics'))
        self.assertEqual(response.status_code, 200)

    def test_admin_users_page_excludes_admin_self(self):
        user = User.objects.create_user(username='regular', password='testpass', is_membership_paid=True, email_verified=True)
        response = self.client.get(reverse('dashboard:admin_users'))
        self.assertContains(response, 'regular')
        self.assertNotContains(response, '<p class="text-sm font-medium text-gray-800">admin</p>')

    def test_admin_update_user_info(self):
        user = User.objects.create_user(username='target', password='testpass', is_membership_paid=True, email_verified=True)
        url = reverse('dashboard:admin_update_user_info', kwargs={'pk': user.pk})
        response = self.client.post(url, {
            'username': 'updatedtarget',
            'email': 'new@test.com',
            'first_name': 'New',
            'last_name': 'Name',
            'phone': '01700000000',
            'student_id': 'ID-1',
            'department': '',
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('dashboard:admin_user_detail', kwargs={'pk': user.pk}) + '?tab=info')
        user.refresh_from_db()
        self.assertEqual(user.username, 'updatedtarget')
        self.assertEqual(user.email, 'new@test.com')
        self.assertEqual(user.first_name, 'New')

    def test_admin_cannot_change_own_password(self):
        response = self.client.get(reverse('accounts:change_password'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('accounts:profile'))

    def test_admin_suspend_user(self):
        user = User.objects.create_user(username='suspendeduser', password='testpass', is_membership_paid=True, email_verified=True)
        suspend_url = reverse('dashboard:admin_suspend_user', kwargs={'pk': user.pk})
        response = self.client.post(suspend_url)
        self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertTrue(user.is_suspended)

    def test_admin_activate_user(self):
        user = User.objects.create_user(username='activeuser', password='testpass', is_suspended=True, is_membership_paid=True, email_verified=True)
        activate_url = reverse('dashboard:admin_activate_user', kwargs={'pk': user.pk})
        response = self.client.post(activate_url)
        self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertFalse(user.is_suspended)


# ========================
# ADMIN REVENUE
# ========================

class TestAdminRevenue(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin = User.objects.create_superuser(username='admin', email='admin@test.com', password='adminpass')
        self.admin.role = 'admin'
        self.admin.save()
        self.student = User.objects.create_user(username='student1', email='student1@test.com', password='testpass', is_membership_paid=True, email_verified=True)
        self.student2 = User.objects.create_user(username='student2', email='student2@test.com', password='testpass', is_membership_paid=True, email_verified=True)
        from apps.payments.models import Payment
        Payment.objects.create(user=self.student, transaction_id='TXN-COMP', amount=Decimal('100'),
                               payment_type='membership', status='completed')
        Payment.objects.create(user=self.student, transaction_id='TXN-PEND', amount=Decimal('200'),
                               payment_type='membership', status='pending')
        Payment.objects.create(user=self.student2, transaction_id='TXN-FAIL', amount=Decimal('300'),
                               payment_type='membership', status='failed')
        Payment.objects.create(user=self.student2, transaction_id='TXN-CANCEL', amount=Decimal('50'),
                               payment_type='reward', status='cancelled')

    def test_admin_can_access_revenue_page(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'admin_dashboard/revenue.html')

    def test_student_cannot_access_revenue_page(self):
        self.client.login(username='student1', password='testpass')
        response = self.client.get(reverse('dashboard:admin_revenue'))
        self.assertEqual(response.status_code, 302)

    def test_anonymous_cannot_access_revenue_page(self):
        response = self.client.get(reverse('dashboard:admin_revenue'))
        self.assertEqual(response.status_code, 302)

    def test_revenue_only_counts_completed_payments(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue'))
        stats = response.context['stats']
        self.assertEqual(stats['total_revenue'], Decimal('100'))
        self.assertEqual(stats['month_revenue'], Decimal('100'))
        self.assertEqual(stats['year_revenue'], Decimal('100'))
        self.assertEqual(stats['successful_payments'], 1)
        self.assertEqual(stats['pending_payments'], 1)
        self.assertEqual(stats['failed_payments'], 1)

    def test_revenue_status_filter(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue') + '?status=completed')
        self.assertEqual(response.context['total_count'], 1)
        payments = list(response.context['payments'])
        self.assertEqual(len(payments), 1)
        self.assertEqual(payments[0].status, 'completed')
        stats = response.context['stats']
        self.assertEqual(stats['total_revenue'], Decimal('100'))

    def test_revenue_payment_type_filter(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue') + '?payment_type=reward')
        self.assertEqual(response.context['total_count'], 1)
        payments = list(response.context['payments'])
        self.assertEqual(payments[0].payment_type, 'reward')

    def test_revenue_search(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue') + '?q=TXN-COMP')
        self.assertEqual(response.context['total_count'], 1)
        response = self.client.get(reverse('dashboard:admin_revenue') + '?q=student2')
        self.assertEqual(response.context['total_count'], 2)
        response = self.client.get(reverse('dashboard:admin_revenue') + '?q=student1@test.com')
        self.assertEqual(response.context['total_count'], 2)

    def test_revenue_custom_date_filter(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue') + '?period=custom&date_from=2000-01-01&date_to=2000-01-02')
        self.assertEqual(response.context['total_count'], 0)
        response = self.client.get(reverse('dashboard:admin_revenue') + '?period=custom&date_from=2000-01-01')
        self.assertEqual(response.context['total_count'], 4)

    def test_revenue_chart_data(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue') + '?chart=month')
        values = json.loads(response.context['chart_values'])
        self.assertEqual(sum(values), 100)

    def test_revenue_export_all(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue_export'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertIn('attachment; filename="payment_history_', response['Content-Disposition'])
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, 'Payment History')
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[:4], ['Transaction ID', 'User Name', 'User Email', 'Payment Type'])
        self.assertIn('SSLCommerz Transaction ID', headers)
        self.assertEqual(ws.max_row, 5)
        self.assertEqual(ws.freeze_panes, 'A2')
        self.assertIsNotNone(ws.auto_filter.ref)
        self.assertEqual(ws['E3'].number_format, '"BDT "#,##0.00')

    def test_revenue_export_respects_filters(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue_export') + '?status=completed')
        from openpyxl import load_workbook
        from io import BytesIO
        ws = load_workbook(BytesIO(response.content)).active
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws['E2'].value, 100)
        response = self.client.get(reverse('dashboard:admin_revenue_export') + '?payment_type=reward')
        ws = load_workbook(BytesIO(response.content)).active
        self.assertEqual(ws.max_row, 2)

    def test_revenue_export_excludes_sensitive_data(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_revenue_export'))
        self.assertNotIn(b'store_passwd', response.content)
        self.assertNotIn(b'SSLCOMMERZ_STORE_PASS', response.content)
        page = self.client.get(reverse('dashboard:admin_revenue')).content
        self.assertNotIn(b'store_passwd', page)


# ========================
# API TESTS
# ========================

class TestAPIs(TestCase):
    def setUp(self):
        self.client = Client()
        from apps.posts.models import Post, Category
        self.user = User.objects.create_user(username='apiuser', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='apiuser', password='testpass')
        self.cat = Category.objects.create(name='TestCat', slug='testcat')
        Post.objects.create(
            user=self.user, title='API Post', description='Test',
            post_type='lost', category=self.cat,
            date_lost_found=timezone.now().date(),
        )

    def test_api_posts_returns_json(self):
        response = self.client.get(reverse('posts:api_posts'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/json')

    def test_api_posts_filters_by_type(self):
        response = self.client.get(reverse('posts:api_posts') + '?type=lost')
        self.assertEqual(response.status_code, 200)
        import json
        data = json.loads(response.content)
        self.assertEqual(len(data['posts']), 1)


# ========================
# NOTIFICATIONS
# ========================

class TestNotifications(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='notifuser', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='notifuser', password='testpass')
        from apps.notifications.models import Notification
        Notification.objects.create(
            user=self.user, notification_type='system',
            title='Test Notification', message='This is a test'
        )

    def test_notification_list_loads(self):
        response = self.client.get(reverse('notifications:list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Test Notification')

    def test_mark_all_read_works(self):
        response = self.client.post(reverse('notifications:mark_all_read'))
        self.assertEqual(response.status_code, 302)
        from apps.notifications.models import Notification
        self.assertEqual(Notification.objects.filter(is_read=False).count(), 0)


# ========================
# SUCCESS STORIES
# ========================

class TestSuccessStories(TestCase):
    def setUp(self):
        self.client = Client()
        from apps.posts.models import Post, SuccessStory
        self.user = User.objects.create_user(username='storyuser', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='storyuser', password='testpass')
        post = Post.objects.create(
            user=self.user, title='Story Post', description='Test',
            post_type='found', date_lost_found=timezone.now().date()
        )
        SuccessStory.objects.create(
            post=post, title='Great Story', story='Someone found an item!',
            finder_name='Finder', owner_name='Owner', is_published=True
        )

    def test_success_stories_page_loads(self):
        response = self.client.get(reverse('posts:success_stories'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Great Story')


# ========================
# TRUST & SAFETY
# ========================

class TestTrustSafety(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='reporter', password='testpass', is_membership_paid=True, email_verified=True)
        self.client.login(username='reporter', password='testpass')
        from apps.posts.models import Post
        self.post = Post.objects.create(
            user=self.user, title='Reportable Post', description='Test',
            post_type='lost', date_lost_found=timezone.now().date()
        )

    def test_report_page_loads(self):
        response = self.client.get(reverse('posts:report_item', kwargs={'post_id': self.post.pk}))
        self.assertEqual(response.status_code, 200)

    def test_report_submission_works(self):
        response = self.client.post(
            reverse('posts:report_item', kwargs={'post_id': self.post.pk}),
            {'report_type': 'spam', 'description': 'This is spam'}
        )
        self.assertEqual(response.status_code, 302)
        from apps.posts.models import TrustReport
        self.assertTrue(TrustReport.objects.filter(post=self.post).exists())


# ========================
# RECOVERY SYSTEM
# ========================

class TestRecoverySystem(TestCase):
    def setUp(self):
        self.client = Client()
        self.owner = User.objects.create_user(username='owner', password='testpass', is_membership_paid=True, email_verified=True)
        self.finder = User.objects.create_user(username='finder', password='testpass', is_membership_paid=True, email_verified=True)
        from apps.posts.models import Post
        self.post = Post.objects.create(
            user=self.owner, title='Recoverable Item', description='Test',
            post_type='lost', date_lost_found=timezone.now().date()
        )
        from apps.recovery.models import RecoverySession
        self.session = RecoverySession.objects.create(
            post=self.post, owner=self.owner, claimant=None,
            status='token_generated',
        )

    def test_recovery_list_requires_login(self):
        response = self.client.get(reverse('recovery:list'))
        self.assertEqual(response.status_code, 302)

    def test_recovery_list_works(self):
        self.client.login(username='owner', password='testpass')
        response = self.client.get(reverse('recovery:list'))
        self.assertEqual(response.status_code, 200)

    def test_recovery_detail_requires_access(self):
        self.client.login(username='finder', password='testpass')
        response = self.client.get(reverse('recovery:detail', args=[self.session.short_code]))
        self.assertEqual(response.status_code, 302)

    def test_owner_can_view_detail(self):
        self.client.login(username='owner', password='testpass')
        response = self.client.get(reverse('recovery:detail', args=[self.session.short_code]))
        self.assertEqual(response.status_code, 200)

    def test_enter_token_requires_finder(self):
        self.client.login(username='owner', password='testpass')
        response = self.client.get(reverse('recovery:enter_token', args=[self.session.short_code]))
        self.assertEqual(response.status_code, 302)

    def test_finder_can_enter_token(self):
        self.session.claimant = self.finder
        self.session.save(update_fields=['claimant'])
        self.client.login(username='finder', password='testpass')
        response = self.client.get(reverse('recovery:enter_token', args=[self.session.short_code]))
        self.assertEqual(response.status_code, 200)

    def test_token_completes_recovery(self):
        self.session.claimant = self.finder
        self.session.save(update_fields=['claimant'])
        self.client.login(username='finder', password='testpass')
        response = self.client.post(reverse('recovery:enter_token', args=[self.session.short_code]),
                                    {'short_code': self.session.short_code})
        self.assertEqual(response.status_code, 302)
        self.session.refresh_from_db()
        self.assertEqual(self.session.status, 'completed')
        self.post.refresh_from_db()
        self.assertTrue(self.post.is_resolved)

    def test_wrong_token_fails(self):
        self.session.claimant = self.finder
        self.session.save(update_fields=['claimant'])
        self.client.login(username='finder', password='testpass')
        response = self.client.post(reverse('recovery:enter_token', args=[self.session.short_code]),
                                    {'short_code': 'WRONG-CODE'})
        self.assertEqual(response.status_code, 302)
        self.session.refresh_from_db()
        self.assertEqual(self.session.status, 'token_generated')

    def test_cancel_session(self):
        self.client.login(username='owner', password='testpass')
        response = self.client.post(reverse('recovery:cancel', args=[self.session.short_code]))
        self.assertEqual(response.status_code, 302)
        self.session.refresh_from_db()
        self.assertEqual(self.session.status, 'cancelled')

    def test_regenerate_token_owner_only(self):
        self.client.login(username='finder', password='testpass')
        response = self.client.post(reverse('recovery:regenerate_token', args=[self.session.short_code]))
        self.assertEqual(response.status_code, 302)

    def test_post_creates_recovery_session(self):
        from apps.posts.models import Post
        from apps.membership.models import Membership, MembershipPlan
        plan = MembershipPlan.objects.create(name='Test Plan', price=100, duration_days=365)
        Membership.objects.create(user=self.owner, plan=plan, is_active=True,
                                  started_at=timezone.now(),
                                  expires_at=timezone.now() + timedelta(days=365))
        self.client.login(username='owner', password='testpass')
        response = self.client.post(reverse('posts:create'), {
            'post_type': 'lost',
            'title': 'Lost Keys',
            'description': 'Lost my keys',
            'date_lost_found': timezone.now().date(),
            'contact_info': 'test@test.com',
        })
        self.assertEqual(response.status_code, 302)
        post = Post.objects.filter(title='Lost Keys').first()
        self.assertIsNotNone(post)
        from apps.recovery.models import RecoverySession
        self.assertTrue(RecoverySession.objects.filter(post=post).exists())


# ========================
# MEMBERSHIP GATING
# ========================

class TestMembershipGating(TestCase):
    def setUp(self):
        self.client = Client()
        self.member = User.objects.create_user(username='member', password='testpass', email='member@test.com', is_membership_paid=True, email_verified=True)
        self.non_member = User.objects.create_user(username='nonmember', password='testpass', email='nonmember@test.com', is_membership_paid=True, email_verified=True)
        self.admin = User.objects.create_superuser(username='admin', password='testpass', email='admin@test.com')
        self.admin.role = 'admin'
        self.admin.save()
        from apps.membership.models import Membership, MembershipPlan
        plan = MembershipPlan.objects.create(name='Annual Membership', price=100, duration_days=365)
        Membership.objects.create(user=self.member, plan=plan, is_active=True,
                                  started_at=timezone.now(),
                                  expires_at=timezone.now() + timedelta(days=365))
        from apps.posts.models import Post, Category, CampusLocation
        self.category = Category.objects.create(name='Electronics', slug='electronics')
        self.location = CampusLocation.objects.create(name='Library', slug='library')
        self.post = Post.objects.create(
            user=self.member, title='Test Item', description='Test',
            post_type='lost', category=self.category, location=self.location,
            date_lost_found=timezone.now().date()
        )

    def test_non_member_cannot_create_post(self):
        self.client.login(username='nonmember', password='testpass')
        response = self.client.get(reverse('posts:create'))
        self.assertEqual(response.status_code, 302)

    def test_member_can_create_post(self):
        self.client.login(username='member', password='testpass')
        response = self.client.get(reverse('posts:create'))
        self.assertEqual(response.status_code, 200)

    def test_admin_cannot_create_post(self):
        self.client.login(username='admin', password='testpass')
        response = self.client.get(reverse('posts:create'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('dashboard:admin_home'))

    def test_non_member_gets_limited_post_detail(self):
        self.client.login(username='nonmember', password='testpass')
        response = self.client.get(reverse('posts:detail', args=[self.post.pk]))
        self.assertContains(response, 'Full Details Hidden')

    def test_member_sees_full_post_detail(self):
        self.client.login(username='member', password='testpass')
        response = self.client.get(reverse('posts:detail', args=[self.post.pk]))
        self.assertNotContains(response, 'Full Details Hidden')


# ========================
# DIRECT MESSAGING
# ========================

class TestMessaging(TestCase):
    def setUp(self):
        self.client = Client()
        self.user1 = User.objects.create_user(username='user1', password='testpass', email='user1@test.com', is_membership_paid=True, email_verified=True)
        self.user2 = User.objects.create_user(username='user2', password='testpass', email='user2@test.com', is_membership_paid=True, email_verified=True)
        from apps.posts.models import Post, Category, CampusLocation
        from apps.membership.models import Membership, MembershipPlan
        self.category = Category.objects.create(name='Electronics', slug='electronics')
        self.location = CampusLocation.objects.create(name='Library', slug='library')
        self.post = Post.objects.create(
            user=self.user1, title='Test Item', description='Test',
            post_type='lost', category=self.category, location=self.location,
            date_lost_found=timezone.now().date()
        )
        plan = MembershipPlan.objects.create(name='Annual Membership', price=100, duration_days=365)
        Membership.objects.create(user=self.user2, plan=plan, is_active=True,
                                  started_at=timezone.now(),
                                  expires_at=timezone.now() + timedelta(days=365))
        Membership.objects.create(user=self.user1, plan=plan, is_active=True,
                                  started_at=timezone.now(),
                                  expires_at=timezone.now() + timedelta(days=365))

    def test_inbox_requires_login(self):
        response = self.client.get(reverse('messaging:inbox'))
        self.assertNotEqual(response.status_code, 200)

    def test_inbox_shows_no_conversations(self):
        self.client.login(username='user1', password='testpass')
        response = self.client.get(reverse('messaging:inbox'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'No conversations yet')

    def test_start_conversation_creates_conv(self):
        self.client.login(username='user2', password='testpass')
        response = self.client.get(reverse('messaging:start', args=[self.post.pk, self.user1.pk]))
        self.assertEqual(response.status_code, 302)
        from apps.messaging.models import Conversation
        self.assertTrue(Conversation.objects.filter(participants=self.user1).filter(participants=self.user2).exists())

    def test_send_message(self):
        self.client.login(username='user2', password='testpass')
        response = self.client.get(reverse('messaging:start', args=[self.post.pk, self.user1.pk]))
        from apps.messaging.models import Conversation
        conv = Conversation.objects.filter(participants=self.user1).filter(participants=self.user2).first()
        response = self.client.post(reverse('messaging:detail', args=[conv.pk]), {'body': 'Hello!'})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(conv.messages.count(), 1)
        self.assertEqual(conv.messages.first().body, 'Hello!')

    def test_self_conversation_not_allowed(self):
        self.client.login(username='user1', password='testpass')
        response = self.client.get(reverse('messaging:start', args=[self.post.pk, self.user1.pk]))
        msgs = list(response.wsgi_request._messages)
        self.assertTrue(any('yourself' in str(m).lower() for m in msgs))

    def test_messages_marked_read(self):
        self.client.login(username='user2', password='testpass')
        self.client.get(reverse('messaging:start', args=[self.post.pk, self.user1.pk]))
        from apps.messaging.models import Conversation, Message
        conv = Conversation.objects.filter(participants=self.user1).filter(participants=self.user2).first()
        Message.objects.create(conversation=conv, sender=self.user1, body='Test message')
        self.client.login(username='user2', password='testpass')
        response = self.client.get(reverse('messaging:detail', args=[conv.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(conv.messages.filter(is_read=False).exclude(sender=self.user2).count(), 0)


# ========================
# TWO-STEP REGISTRATION
# ========================

class TestTwoStepRegistration(TestCase):
    def setUp(self):
        self.client = Client()
        self.register_url = reverse('accounts:register')
        self.login_url = reverse('accounts:login')
        from apps.membership.models import MembershipPlan
        self.plan = MembershipPlan.objects.create(
            name='Annual', price=Decimal('500.00'), duration_days=365, is_active=True
        )

    def test_registration_creates_pending_user(self):
        data = {
            'username': 'newstudent',
            'email': 'new@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'student_id': '2021-001-002',
            'department': 'cse',
        }
        response = self.client.post(self.register_url, data)
        self.assertEqual(response.status_code, 302)
        user = User.objects.get(username='newstudent')
        self.assertFalse(user.is_membership_paid)
        self.assertEqual(user.role, 'student')

    def test_registration_auto_logins_and_redirects_to_verify_email(self):
        data = {
            'username': 'autoLogin',
            'email': 'auto@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'bba',
        }
        response = self.client.post(self.register_url, data)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/verify-email/', response.url)

    def test_pending_user_sees_membership_purchase_page(self):
        data = {
            'username': 'pendinguser',
            'email': 'pending@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='pendinguser')
        user.email_verified = True
        user.save()
        response = self.client.get(reverse('membership:pending_purchase'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Complete Your Registration')
        self.assertContains(response, 'Membership Payment')

    def test_pending_user_blocked_from_dashboard(self):
        data = {
            'username': 'blockeduser',
            'email': 'blocked@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='blockeduser')
        user.email_verified = True
        user.save()
        response = self.client.get(reverse('dashboard:home'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/membership/pending/', response.url)

    def test_pending_user_blocked_from_browse(self):
        data = {
            'username': 'blockeduser2',
            'email': 'blocked2@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='blockeduser2')
        user.email_verified = True
        user.save()
        response = self.client.get(reverse('posts:browse'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/membership/pending/', response.url)

    def test_pending_user_can_access_membership_pages(self):
        data = {
            'username': 'canaccess',
            'email': 'access@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='canaccess')
        user.email_verified = True
        user.save()
        response = self.client.get(reverse('membership:pending_purchase'))
        self.assertEqual(response.status_code, 200)
        response = self.client.get(reverse('membership:index'))
        self.assertEqual(response.status_code, 200)

    def test_pending_user_can_logout(self):
        data = {
            'username': 'canlogout',
            'email': 'logout@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='canlogout')
        user.email_verified = True
        user.save()
        response = self.client.get(reverse('accounts:logout'))
        self.assertEqual(response.status_code, 302)

    def test_payment_success_activates_user(self):
        from apps.payments.models import Payment
        data = {
            'username': 'payuser',
            'email': 'pay@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='payuser')
        self.assertFalse(user.is_membership_paid)
        payment = Payment.objects.create(
            user=user, amount=Decimal('500.00'), payment_type='membership',
            status='pending', sslcommerz_tran_id='TEST-TRAN-001',
        )
        with patch('apps.payments.views.verify_sslcommerz_payment') as mock_verify:
            mock_verify.return_value = {'status': 'VALID', 'bank_tran_id': 'BANK001'}
            response = self.client.post(reverse('payments:success'), {
                'tran_id': 'TEST-TRAN-001', 'val_id': 'VAL001',
            })
            self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertTrue(user.is_membership_paid)

    def test_returning_user_redirected_to_membership(self):
        user = User.objects.create_user(
            username='returning', password='testpass', email='return@test.com',
            is_membership_paid=False, email_verified=True,
        )
        self.client.login(username='returning', password='testpass')
        response = self.client.get(reverse('dashboard:home'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/membership/pending/', response.url)

    def test_paid_user_accesses_dashboard_normally(self):
        user = User.objects.create_user(
            username='paiduser', password='testpass', email='paid@test.com',
            is_membership_paid=True, email_verified=True,
        )
        self.client.login(username='paiduser', password='testpass')
        response = self.client.get(reverse('dashboard:home'))
        self.assertEqual(response.status_code, 200)

    def test_admin_bypasses_membership_check(self):
        admin = User.objects.create_superuser(username='admin2', email='a@t.com', password='adminpass')
        admin.role = 'admin'
        admin.save()
        self.client.login(username='admin2', password='adminpass')
        response = self.client.get(reverse('dashboard:admin_home'))
        self.assertEqual(response.status_code, 200)

    def test_duplicate_payment_not_doubled(self):
        from apps.payments.models import Payment
        data = {
            'username': 'dupuser',
            'email': 'dup@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='dupuser')
        payment = Payment.objects.create(
            user=user, amount=Decimal('500.00'), payment_type='membership',
            status='completed', sslcommerz_tran_id='COMPLETED-001',
            transaction_id='BANK-COMPLETE',
        )
        with patch('apps.payments.views.verify_sslcommerz_payment') as mock_verify:
            mock_verify.return_value = {'status': 'VALID', 'bank_tran_id': 'BANK001'}
            response = self.client.post(reverse('payments:success'), {
                'tran_id': 'COMPLETED-001', 'val_id': 'VAL001',
            })
            self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertTrue(user.is_membership_paid)

    def test_failed_payment_keeps_user_pending(self):
        from apps.payments.models import Payment
        data = {
            'username': 'failuser',
            'email': 'fail@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='failuser')
        payment = Payment.objects.create(
            user=user, amount=Decimal('500.00'), payment_type='membership',
            status='pending', sslcommerz_tran_id='FAIL-TRAN-001',
        )
        response = self.client.post(reverse('payments:fail'), {
            'tran_id': 'FAIL-TRAN-001',
        })
        self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertFalse(user.is_membership_paid)
        payment.refresh_from_db()
        self.assertEqual(payment.status, 'failed')

    def test_cancelled_payment_keeps_user_pending(self):
        from apps.payments.models import Payment
        data = {
            'username': 'canceluser',
            'email': 'cancel@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='canceluser')
        payment = Payment.objects.create(
            user=user, amount=Decimal('500.00'), payment_type='membership',
            status='pending', sslcommerz_tran_id='CANCEL-TRAN-001',
        )
        response = self.client.post(reverse('payments:cancel'), {
            'tran_id': 'CANCEL-TRAN-001',
        })
        self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertFalse(user.is_membership_paid)
        payment.refresh_from_db()
        self.assertEqual(payment.status, 'cancelled')

    def test_pending_user_can_retry_payment(self):
        from apps.payments.models import Payment
        data = {
            'username': 'retryuser',
            'email': 'retry@student.com',
            'password1': 'ComplexPass123!',
            'password2': 'ComplexPass123!',
            'department': 'cse',
        }
        self.client.post(self.register_url, data)
        user = User.objects.get(username='retryuser')
        user.email_verified = True
        user.save()
        Payment.objects.create(
            user=user, amount=Decimal('500.00'), payment_type='membership',
            status='failed', sslcommerz_tran_id='RETRY-001',
        )
        response = self.client.get(reverse('membership:pending_purchase'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Purchase Membership')