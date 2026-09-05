import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
from django.db import connection
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import timedelta, datetime, time
import json

from .models import User, UserActivity
from apps.posts.models import Post, Category, CampusLocation
from apps.membership.models import Membership, MembershipPlan
from apps.payments.models import Payment
from apps.ai_engine.models import PostEmbedding, MatchSuggestion
from apps.notifications.models import Notification

logger = logging.getLogger(__name__)


def is_admin(user):
    return user.is_authenticated and (user.role == 'admin' or user.is_staff or user.is_superuser)


@login_required
def user_dashboard(request):
    user = request.user
    posts = Post.objects.select_related('location', 'category').filter(user=user).order_by('-created_at')
    recent_activities = UserActivity.objects.filter(user=user)[:10]
    membership = getattr(user, 'membership', None)
    matches = MatchSuggestion.objects.select_related('post', 'matched_post').filter(post__user=user)[:5]
    unread_notifications = Notification.objects.filter(user=user, is_read=False).count()
    site_agg = Post.objects.aggregate(
        total_posts=Count('id'),
        open_posts=Count('id', filter=Q(status='open')),
        resolved_posts=Count('id', filter=Q(status='resolved')),
        lost_posts=Count('id', filter=Q(post_type='lost')),
        found_posts=Count('id', filter=Q(post_type='found')),
    )

    context = {
        'total_posts': site_agg['total_posts'],
        'open_posts': site_agg['open_posts'],
        'resolved_posts': site_agg['resolved_posts'],
        'lost_posts': site_agg['lost_posts'],
        'found_posts': site_agg['found_posts'],
        'unread_notifications': unread_notifications,
        'posts': posts,
        'recent_activities': recent_activities,
        'membership': membership,
        'matches': matches,
    }
    return render(request, 'dashboard/user_dashboard.html', context)


ADMIN_SIDEBAR = [
    {'url_name': 'dashboard:admin_home', 'label': 'Dashboard', 'icon': 'bi bi-grid-1x2'},
    {'url_name': 'dashboard:admin_users', 'label': 'Users', 'icon': 'bi bi-people'},
    {'url_name': 'posts:browse', 'label': 'Browse Posts', 'icon': 'bi bi-collection'},
    {'url_name': 'notifications:list', 'label': 'Notifications', 'icon': 'bi bi-bell'},
    {'url_name': 'dashboard:admin_memberships', 'label': 'Memberships', 'icon': 'bi bi-gem'},
    {'url_name': 'dashboard:admin_revenue', 'label': 'Revenue', 'icon': 'bi bi-currency-dollar'},
    {'url_name': 'dashboard:admin_settings', 'label': 'Settings', 'icon': 'bi bi-gear'},
]


@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    users_agg = User.objects.aggregate(
        total_users=Count('id'),
        active_users=Count('id', filter=Q(is_active=True, is_suspended=False)),
    )
    total_users = users_agg['total_users']
    active_users = users_agg['active_users']
    inactive_users = total_users - active_users
    posts_agg = Post.objects.aggregate(
        total_posts=Count('id'),
        open_posts=Count('id', filter=Q(status='open')),
        resolved_posts=Count('id', filter=Q(status='resolved')),
        lost_posts=Count('id', filter=Q(post_type='lost')),
        found_posts=Count('id', filter=Q(post_type='found')),
    )
    total_posts = posts_agg['total_posts']
    open_posts = posts_agg['open_posts']
    resolved_posts = posts_agg['resolved_posts']
    lost_posts = posts_agg['lost_posts']
    found_posts = posts_agg['found_posts']
    payments_agg = Payment.objects.aggregate(
        total_revenue=Sum('amount', filter=Q(status='completed')),
        pending_payments=Count('id', filter=Q(status='pending')),
    )
    total_revenue = payments_agg['total_revenue'] or 0
    pending_payments = payments_agg['pending_payments']

    posts_by_category = list(Post.objects.values('category__name').annotate(count=Count('id')))
    posts_by_location = list(Post.objects.values('location__name').annotate(count=Count('id')))
    try:
        monthly_registrations = list(User.objects.annotate(month=TruncMonth('date_joined')).values('month').annotate(count=Count('id')).order_by('-month')[:12])
    except Exception:
        monthly_registrations = []
    try:
        monthly_revenue = list(Payment.objects.filter(status='completed').annotate(month=TruncMonth('created_at')).values('month').annotate(total=Sum('amount')).order_by('-month')[:12])
    except Exception:
        monthly_revenue = []
    from apps.recovery.models import RecoverySession
    recovery_agg = RecoverySession.objects.aggregate(
        recovery_sessions=Count('id'),
        completed_recoveries=Count('id', filter=Q(status='completed')),
        pending_recoveries=Count('id', filter=Q(status__in=['pending', 'token_generated'])),
    )
    recent_activities = UserActivity.objects.all()[:20]
    users = User.objects.exclude(pk=request.user.pk).order_by('-date_joined')[:10]
    posts = Post.objects.select_related('location', 'category').order_by('-created_at')[:10]
    payments = Payment.objects.select_related('user').all().order_by('-created_at')[:10]

    context = {
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'total_posts': total_posts,
        'open_posts': open_posts,
        'resolved_posts': resolved_posts,
        'lost_posts': lost_posts,
        'found_posts': found_posts,
        'total_revenue': total_revenue,
        'pending_payments': pending_payments,
        'recovery_sessions': recovery_agg['recovery_sessions'],
        'completed_recoveries': recovery_agg['completed_recoveries'],
        'pending_recoveries': recovery_agg['pending_recoveries'],
        'posts_by_category': posts_by_category,
        'posts_by_location': posts_by_location,
        'monthly_registrations': monthly_registrations,
        'monthly_revenue': monthly_revenue,
        'recent_activities': recent_activities,
        'users': users,
        'posts': posts,
        'payments': payments,
        'sidebar_items': ADMIN_SIDEBAR,
    }
    return render(request, 'admin_dashboard/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def admin_users(request):
    users = User.objects.exclude(pk=request.user.pk).order_by('-date_joined')
    paginator = Paginator(users, 20)
    page = request.GET.get('page', 1)
    users_page = paginator.get_page(page)
    return render(request, 'admin_dashboard/users.html', {'users': users_page, 'sidebar_items': ADMIN_SIDEBAR})


@login_required
@user_passes_test(is_admin)
def admin_user_detail(request, pk):
    from apps.payments.models import Payment
    from apps.membership.models import MembershipPlan
    user = get_object_or_404(User, pk=pk)
    user_posts = Post.objects.filter(user=user).select_related('category', 'location')
    user_payments = Payment.objects.filter(user=user).select_related('user')
    user_activities = UserActivity.objects.filter(user=user)[:20]
    open_posts = user_posts.filter(status='open').count()
    resolved_posts = user_posts.filter(status='resolved').count()
    payment_count = user_payments.count()
    return render(request, 'admin_dashboard/user_detail.html', {
        'profile_user': user,
        'user_posts': user_posts,
        'user_payments': user_payments,
        'user_activities': user_activities,
        'posts': user_posts,
        'activities': user_activities,
        'open_posts': open_posts,
        'resolved_posts': resolved_posts,
        'payment_count': payment_count,
        'total_posts': user_posts.count(),
        'plans': MembershipPlan.objects.filter(is_active=True),
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_suspend_user(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_users')
    user = get_object_or_404(User, pk=pk)
    user.is_suspended = True
    user.is_active = False
    user.save()
    logger.info('Admin %s suspended user %s', request.user.username, user.username)
    messages.success(request, f'User {user.username} has been suspended.')
    return redirect('dashboard:admin_users')


@login_required
@user_passes_test(is_admin)
def admin_activate_user(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_users')
    user = get_object_or_404(User, pk=pk)
    user.is_suspended = False
    user.is_active = True
    user.save()
    logger.info('Admin %s activated user %s', request.user.username, user.username)
    messages.success(request, f'User {user.username} has been activated.')
    return redirect('dashboard:admin_users')


@login_required
@user_passes_test(is_admin)
def admin_posts(request):
    posts = Post.objects.select_related('user', 'category', 'location').all().order_by('-created_at')
    paginator = Paginator(posts, 20)
    page = request.GET.get('page', 1)
    posts_page = paginator.get_page(page)
    return render(request, 'admin_dashboard/posts.html', {
        'posts': posts_page,
        'total_posts_count': posts.count(),
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_post_approve(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_posts')
    post = get_object_or_404(Post, pk=pk)
    post.status = 'open'
    post.save(update_fields=['status'])
    messages.success(request, f'Post "{post.title}" approved.')
    return redirect('dashboard:admin_posts')


@login_required
@user_passes_test(is_admin)
def admin_post_reject(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_posts')
    post = get_object_or_404(Post, pk=pk)
    post.status = 'claimed'
    post.save(update_fields=['status'])
    messages.success(request, f'Post "{post.title}" rejected.')
    return redirect('dashboard:admin_posts')


@login_required
@user_passes_test(is_admin)
def admin_post_soft_delete(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_posts')
    post = get_object_or_404(Post, pk=pk)
    post.is_active = False
    post.save(update_fields=['is_active'])
    messages.success(request, f'Post "{post.title}" soft deleted.')
    return redirect('dashboard:admin_posts')


@login_required
@user_passes_test(is_admin)
def admin_post_restore(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_posts')
    post = get_object_or_404(Post, pk=pk)
    post.is_active = True
    post.save(update_fields=['is_active'])
    messages.success(request, f'Post "{post.title}" restored.')
    return redirect('dashboard:admin_posts')


@login_required
@user_passes_test(is_admin)
def admin_post_permanent_delete(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_posts')
    post = get_object_or_404(Post, pk=pk)
    title = post.title
    post.delete()
    messages.success(request, f'Post "{title}" permanently deleted.')
    return redirect('dashboard:admin_posts')


@login_required
@user_passes_test(is_admin)
def admin_categories(request):
    categories = Category.objects.annotate(post_count=Count('posts'))
    total_categories = categories.count()
    active_categories = categories.filter(is_active=True).count()
    total_posts_in_categories = Post.objects.filter(category__isnull=False).count()
    return render(request, 'admin_dashboard/categories.html', {
        'categories': categories,
        'total_categories': total_categories,
        'active_categories': active_categories,
        'total_posts_in_categories': total_posts_in_categories,
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_locations(request):
    locations = CampusLocation.objects.annotate(post_count=Count('posts'))
    total_locations = locations.count()
    active_locations = locations.filter(is_active=True).count()
    total_posts_in_locations = Post.objects.filter(location__isnull=False).count()
    return render(request, 'admin_dashboard/locations.html', {
        'locations': locations,
        'total_locations': total_locations,
        'active_locations': active_locations,
        'total_posts_in_locations': total_posts_in_locations,
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_revenue(request):
    qs, params = _revenue_filters(request)

    now = timezone.localtime()
    tz = timezone.get_current_timezone()
    today = now.date()
    month_start = today.replace(day=1)
    next_month = (month_start + timedelta(days=32)).replace(day=1)
    year_start = today.replace(month=1, day=1)
    next_year = year_start.replace(year=year_start.year + 1)

    def local_midnight(day):
        return timezone.make_aware(datetime.combine(day, time.min), tz)

    stats = qs.aggregate(
        total_revenue=Sum('amount', filter=Q(status='completed')),
        month_revenue=Sum('amount', filter=Q(
            status='completed',
            created_at__gte=local_midnight(month_start),
            created_at__lt=local_midnight(next_month),
        )),
        year_revenue=Sum('amount', filter=Q(
            status='completed',
            created_at__gte=local_midnight(year_start),
            created_at__lt=local_midnight(next_year),
        )),
        successful_payments=Count('id', filter=Q(status='completed')),
        pending_payments=Count('id', filter=Q(status='pending')),
        failed_payments=Count('id', filter=Q(status='failed')),
    )

    chart_labels, chart_values = _revenue_chart_series(
        qs, params['chart'], params['date_from'], params['date_to']
    )

    plan_names = _plan_name_map()
    total_count = qs.count()
    paginator = Paginator(qs, 20)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    for p in page_obj.object_list:
        p.plan_name = plan_names.get(str(p.reference_id), '—')
        p.detail_json = json.dumps({
            'transaction_id': p.transaction_id or 'N/A',
            'user_name': p.user.get_full_name() or p.user.username,
            'email': p.user.email or '—',
            'membership': p.plan_name,
            'amount': f'{p.amount:,.2f} BDT',
            'status': p.get_status_display(),
            'gateway': p.gateway,
            'method': p.payment_method,
            'sslcommerz_tran_id': p.sslcommerz_tran_id or 'N/A',
            'payment_date': timezone.localtime(p.created_at).strftime('%b %d, %Y, %I:%M %p'),
            'completion_date': timezone.localtime(p.completion_date).strftime('%b %d, %Y, %I:%M %p') if p.completion_date else '—',
        })

    from urllib.parse import urlencode
    filter_qs = urlencode({k: v for k, v in params.items() if v})

    return render(request, 'admin_dashboard/revenue.html', {
        **params,
        'stats': stats,
        'total_count': total_count,
        'payments': page_obj,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
        'chart_period': params['chart'],
        'filter_qs': filter_qs,
        'status_choices': Payment.PAYMENT_STATUS,
        'type_choices': Payment.PAYMENT_TYPES,
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_revenue_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    qs, params = _revenue_filters(request)

    headers = [
        'Transaction ID', 'User Name', 'User Email', 'Payment Type', 'Amount',
        'Payment Status', 'Payment Gateway', 'Payment Method',
        'SSLCommerz Transaction ID', 'Payment Date',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Payment History'
    ws.append(headers)

    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    tz = timezone.get_current_timezone()
    for p in qs.iterator():
        ws.append([
            p.transaction_id or 'N/A',
            p.user.get_full_name() or p.user.username,
            p.user.email or '',
            p.get_payment_type_display(),
            float(p.amount),
            p.get_status_display(),
            p.gateway,
            p.payment_method,
            p.sslcommerz_tran_id or 'N/A',
            p.created_at.astimezone(tz).replace(tzinfo=None),
        ])

    for i, width in enumerate([22, 22, 26, 14, 12, 16, 16, 16, 24, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        row[4].number_format = '"BDT "#,##0.00'
        row[9].number_format = 'YYYY-MM-DD HH:MM'

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    today = timezone.localtime().strftime('%Y-%m-%d')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payment_history_{today}.xlsx"'
    wb.save(response)
    return response


def _plan_name_map():
    return {str(plan.pk): plan.name for plan in MembershipPlan.objects.all()}


def _revenue_filters(request):
    """Build the filtered Payment queryset plus the applied filter params."""
    q = request.GET.get('q', '').strip()
    period = request.GET.get('period', 'all')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    status = request.GET.get('status', '')
    payment_type = request.GET.get('payment_type', '')
    chart = request.GET.get('chart', 'month')

    if status not in [s[0] for s in Payment.PAYMENT_STATUS]:
        status = ''
    if payment_type not in [t[0] for t in Payment.PAYMENT_TYPES]:
        payment_type = ''
    if chart not in ('day', 'week', 'month'):
        chart = 'month'

    now = timezone.localtime()
    tz = timezone.get_current_timezone()
    today = now.date()

    def local_midnight(day):
        return timezone.make_aware(datetime.combine(day, time.min), tz)

    qs = Payment.objects.select_related('user').order_by('-created_at')
    if period == 'today':
        qs = qs.filter(
            created_at__gte=local_midnight(today),
            created_at__lt=local_midnight(today + timedelta(days=1)),
        )
    elif period == '7d':
        qs = qs.filter(created_at__gte=now - timedelta(days=7))
    elif period == 'month':
        month_start = today.replace(day=1)
        next_month = (month_start + timedelta(days=32)).replace(day=1)
        qs = qs.filter(
            created_at__gte=local_midnight(month_start),
            created_at__lt=local_midnight(next_month),
        )
    elif period == 'year':
        year_start = today.replace(month=1, day=1)
        next_year = year_start.replace(year=year_start.year + 1)
        qs = qs.filter(
            created_at__gte=local_midnight(year_start),
            created_at__lt=local_midnight(next_year),
        )
    elif period == 'custom':
        try:
            d1 = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        except ValueError:
            d1 = None
        try:
            d2 = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        except ValueError:
            d2 = None
        if d1:
            qs = qs.filter(created_at__gte=local_midnight(d1))
        if d2:
            qs = qs.filter(created_at__lt=local_midnight(d2 + timedelta(days=1)))

    if status:
        qs = qs.filter(status=status)
    if payment_type:
        qs = qs.filter(payment_type=payment_type)

    if q:
        qs = qs.filter(
            Q(transaction_id__icontains=q)
            | Q(sslcommerz_tran_id__icontains=q)
            | Q(user__username__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(user__email__icontains=q)
        )

    return qs, {
        'q': q,
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'payment_type': payment_type,
        'chart': chart,
    }


def _revenue_chart_series(qs, chart, date_from='', date_to=''):
    """Aggregate completed payments into a zero-filled time series for the chart."""
    from decimal import Decimal
    rev = qs.filter(status='completed')
    today = timezone.localtime().date()

    try:
        d_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
    except (ValueError, TypeError):
        d_from = None
    try:
        d_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
    except (ValueError, TypeError):
        d_to = None

    if chart == 'day':
        if d_from:
            start = d_from
        elif d_to:
            start = d_to - timedelta(days=29)
        else:
            start = today - timedelta(days=29)
        end = d_to or today
        if end < start:
            end = start
        days = []
        d = start
        while d <= end and len(days) < 60:
            days.append(d)
            d += timedelta(days=1)
        totals = {d: Decimal('0') for d in days}
        for row in rev.annotate(day=TruncDate('created_at')).values('day').annotate(total=Sum('amount')):
            if row['day'] in totals:
                totals[row['day']] += row['total'] or Decimal('0')
        return [d.strftime('%b %d') for d in days], [float(totals[d]) for d in days]

    if chart == 'week':
        this_week = today - timedelta(days=today.weekday())
        if d_from:
            weeks = []
            w = d_from - timedelta(days=d_from.weekday())
            end = d_to or today
            while w <= end and len(weeks) < 26:
                weeks.append(w)
                w += timedelta(weeks=1)
            if not weeks:
                weeks = [this_week]
        else:
            weeks = [this_week - timedelta(weeks=i) for i in range(11, -1, -1)]
        totals = {w.isocalendar()[:2]: Decimal('0') for w in weeks}
        for row in rev.annotate(week=TruncWeek('created_at')).values('week').annotate(total=Sum('amount')):
            if row['week']:
                key = row['week'].isocalendar()[:2]
                if key in totals:
                    totals[key] += row['total'] or Decimal('0')
        return [w.strftime('%b %d') for w in weeks], [float(totals[w.isocalendar()[:2]]) for w in weeks]

    if d_from:
        months = []
        y, m = d_from.year, d_from.month
        end = d_to or today
        while (y, m) <= (end.year, end.month) and len(months) < 24:
            months.append((y, m))
            m += 1
            if m > 12:
                m = 1
                y += 1
        if not months:
            months = [(today.year, today.month)]
    else:
        months = []
        y, m = today.year, today.month
        for _ in range(12):
            months.append((y, m))
            m -= 1
            if m < 1:
                m = 12
                y -= 1
        months.reverse()

    from datetime import date as _date
    totals = {f'{y}-{m:02d}': Decimal('0') for y, m in months}
    for row in rev.annotate(month=TruncMonth('created_at')).values('month').annotate(total=Sum('amount')):
        if row['month']:
            key = f"{row['month'].year}-{row['month'].month:02d}"
            if key in totals:
                totals[key] += row['total'] or Decimal('0')
    labels = [_date(y, m, 1).strftime('%b %Y') for y, m in months]
    return labels, [float(totals[f'{y}-{m:02d}']) for y, m in months]


@login_required
@user_passes_test(is_admin)
def admin_reports(request):
    total_users = User.objects.count()
    total_posts = Post.objects.count()
    resolved_posts = Post.objects.filter(status='resolved').count()
    recovery_rate = round((resolved_posts / total_posts * 100) if total_posts > 0 else 0, 1)
    posts_by_category = list(Post.objects.values('category__name').annotate(count=Count('id')))
    posts_by_location = list(Post.objects.values('location__name').annotate(count=Count('id')))
    total_revenue = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    return render(request, 'admin_dashboard/reports.html', {
        'total_users': total_users,
        'total_posts': total_posts,
        'resolved_posts': resolved_posts,
        'recovery_rate': recovery_rate,
        'posts_by_category': posts_by_category,
        'posts_by_location': posts_by_location,
        'total_revenue': total_revenue,
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_analytics(request):
    total_users = User.objects.count()
    total_posts = Post.objects.count()
    resolved_posts = Post.objects.filter(status='resolved').count()
    recovery_rate = round((resolved_posts / total_posts * 100) if total_posts > 0 else 0, 1)
    posts_by_category = list(Post.objects.values('category__name').annotate(count=Count('id')))
    posts_by_location = list(Post.objects.values('location__name').annotate(count=Count('id')))
    monthly_registrations = list(User.objects.annotate(month=TruncMonth('date_joined'))
        .values('month').annotate(count=Count('id')).order_by('-month')[:12])
    monthly_revenue = list(Payment.objects.filter(status='completed')
        .annotate(month=TruncMonth('created_at'))
        .values('month').annotate(total=Sum('amount')).order_by('-month')[:12])
    recovery_success_rate = list(Post.objects.filter(status='resolved')
        .annotate(month=TruncMonth('updated_at'))
        .values('month').annotate(count=Count('id')).order_by('-month')[:12])
    return render(request, 'admin_dashboard/analytics.html', {
        'total_users': total_users,
        'total_posts': total_posts,
        'resolved_posts': resolved_posts,
        'recovery_rate': recovery_rate,
        'posts_by_category': posts_by_category,
        'posts_by_location': posts_by_location,
        'monthly_registrations': monthly_registrations,
        'monthly_revenue': monthly_revenue,
        'recovery_success_rate': recovery_success_rate,
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_settings(request):
    return render(request, 'admin_dashboard/settings.html', {'sidebar_items': ADMIN_SIDEBAR})


@login_required
@user_passes_test(is_admin)
def admin_memberships(request):
    memberships = Membership.objects.select_related('user', 'plan').order_by('-created_at')
    total_members = memberships.count()
    active_members = memberships.filter(is_active=True).count()
    expired_members = total_members - active_members
    paginator = Paginator(memberships, 20)
    page = request.GET.get('page', 1)
    memberships_page = paginator.get_page(page)
    return render(request, 'admin_dashboard/memberships.html', {
        'memberships': memberships_page,
        'total_members': total_members,
        'active_members': active_members,
        'expired_members': expired_members,
        'sidebar_items': ADMIN_SIDEBAR,
    })


@login_required
@user_passes_test(is_admin)
def admin_update_membership(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_user_detail', pk=pk)
    user = get_object_or_404(User, pk=pk)
    membership = getattr(user, 'membership', None)

    is_active = request.POST.get('is_active') == 'on'
    plan_id = request.POST.get('plan')
    started_at_str = request.POST.get('started_at')
    expires_at_str = request.POST.get('expires_at')

    plan = MembershipPlan.objects.filter(pk=plan_id).first() if plan_id else None
    if is_active and not plan:
        messages.error(request, 'Please select a plan for active membership.')
        return redirect('dashboard:admin_user_detail', pk=pk)

    started_at = timezone.datetime.strptime(started_at_str, '%Y-%m-%d').replace(tzinfo=timezone.get_current_timezone()) if started_at_str else None
    expires_at = timezone.datetime.strptime(expires_at_str, '%Y-%m-%d').replace(tzinfo=timezone.get_current_timezone()) if expires_at_str else None

    if is_active and (not started_at or not expires_at):
        messages.error(request, 'Start and expiry dates are required for active membership.')
        return redirect('dashboard:admin_user_detail', pk=pk)

    if membership:
        membership.plan = plan
        membership.is_active = is_active
        membership.started_at = started_at
        membership.expires_at = expires_at
        membership.save()
    else:
        membership = Membership.objects.create(
            user=user, plan=plan, is_active=is_active,
            started_at=started_at, expires_at=expires_at
        )
    user.is_membership_paid = is_active
    user.save(update_fields=['is_membership_paid'])
    messages.success(request, f'Membership updated for {user.username}.')
    return redirect('dashboard:admin_user_detail', pk=pk)


@login_required
@user_passes_test(is_admin)
def admin_toggle_membership(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_user_detail', pk=pk)
    user = get_object_or_404(User, pk=pk)
    membership = getattr(user, 'membership', None)
    if membership and membership.is_active:
        membership.is_active = False
        membership.expires_at = timezone.now()
        membership.save()
        user.is_membership_paid = False
        user.save(update_fields=['is_membership_paid'])
        messages.success(request, f'Membership revoked for {user.username}.')
    else:
        plan = MembershipPlan.objects.filter(is_active=True).first()
        if not plan:
            messages.error(request, 'No active membership plan found. Create one first.')
            return redirect('dashboard:admin_user_detail', pk=pk)
        if membership:
            membership.plan = plan
            membership.is_active = True
            membership.started_at = timezone.now()
            membership.expires_at = timezone.now() + timedelta(days=plan.duration_days)
            membership.save()
        else:
            membership = Membership.objects.create(
                user=user, plan=plan, is_active=True,
                started_at=timezone.now(),
                expires_at=timezone.now() + timedelta(days=plan.duration_days)
            )
        user.is_membership_paid = True
        user.save(update_fields=['is_membership_paid'])
        messages.success(request, f'Membership granted to {user.username}.')
    return redirect('dashboard:admin_user_detail', pk=pk)


@login_required
@user_passes_test(is_admin)
def admin_extend_membership(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_user_detail', pk=pk)
    user = get_object_or_404(User, pk=pk)
    try:
        days = int(request.POST.get('days', 30))
        days = max(1, min(days, 3650))
    except (TypeError, ValueError):
        days = 30
    membership = getattr(user, 'membership', None)
    if membership:
        base = membership.expires_at if membership.expires_at and membership.expires_at > timezone.now() else timezone.now()
        membership.expires_at = base + timedelta(days=days)
        membership.is_active = True
        if not membership.started_at:
            membership.started_at = timezone.now()
        membership.save()
    else:
        plan = MembershipPlan.objects.filter(is_active=True).first()
        if not plan:
            messages.error(request, 'No active membership plan found. Create one first.')
            return redirect('dashboard:admin_user_detail', pk=pk)
        Membership.objects.create(
            user=user, plan=plan, is_active=True,
            started_at=timezone.now(),
            expires_at=timezone.now() + timedelta(days=days),
        )
    user.is_membership_paid = True
    user.save(update_fields=['is_membership_paid'])
    messages.success(request, f'Membership extended by {days} days for {user.username}.')
    return redirect('dashboard:admin_user_detail', pk=pk)


@login_required
@user_passes_test(is_admin)
def admin_delete_user(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_users')
    user = get_object_or_404(User, pk=pk)
    if user.pk == request.user.pk:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('dashboard:admin_users')
    username = user.username
    user.delete()
    logger.info('Admin %s deleted user %s', request.user.username, username)
    messages.success(request, f'User "{username}" deleted.')
    return redirect('dashboard:admin_users')


@login_required
@user_passes_test(is_admin)
def admin_update_user_info(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:admin_user_detail', pk=pk)
    user = get_object_or_404(User, pk=pk)
    username = request.POST.get('username', '').strip()
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    email = request.POST.get('email', '').strip()
    phone = request.POST.get('phone', '').strip()
    student_id = request.POST.get('student_id', '').strip()
    department = request.POST.get('department', '').strip()

    valid_departments = [d[0] for d in User.DEPARTMENTS]
    errors = []
    if not username:
        errors.append('Username is required.')
    elif User.objects.exclude(pk=user.pk).filter(username__iexact=username).exists():
        errors.append('That username is already taken.')
    if not email:
        errors.append('Email is required.')
    elif User.objects.exclude(pk=user.pk).filter(email__iexact=email).exists():
        errors.append('That email is already in use.')
    if student_id and User.objects.exclude(pk=user.pk).filter(student_id__iexact=student_id).exists():
        errors.append('That student ID is already in use.')
    if department and department not in valid_departments:
        department = ''

    if errors:
        for error in errors:
            messages.error(request, error)
    else:
        user.username = username
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.phone = phone or None
        user.student_id = student_id or None
        user.department = department or None
        user.save()
        UserActivity.objects.create(
            user=user,
            activity_type='profile_updated',
            description=f'Information updated by admin {request.user.username}.',
        )
        messages.success(request, f'Information updated for {user.username}.')
    return redirect(f"{reverse('dashboard:admin_user_detail', kwargs={'pk': pk})}?tab=info")
